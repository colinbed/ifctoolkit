from __future__ import annotations
import csv
from pathlib import Path
from typing import Iterable
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from backend.cobie_models import CobieValidationResult
ISSUE_FIELDS=["issue_id","severity","sheet_name","row_number","column_name","cell_reference","rule_id","rule_type","message","expected_value","actual_value","recommendation","related_sheet","related_column"]

def autosize(ws):
    for col in range(1, ws.max_column+1):
        letter=get_column_letter(col); ws.column_dimensions[letter].width=min(max(12, max(len(str(ws.cell(r,col).value or "")) for r in range(1, min(ws.max_row,80)+1))+2),60)

def write_csv(result: CobieValidationResult, path: str|Path) -> Path:
    path=Path(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        w=csv.DictWriter(f, fieldnames=ISSUE_FIELDS); w.writeheader()
        for issue in result.issues: w.writerow(issue.dict())
    return path

def write_excel_report(result: CobieValidationResult, path: str|Path, file_size: int = 0, environment: str = "production") -> Path:
    path=Path(path); wb=Workbook(); ws=wb.active; ws.title="Summary"
    rows=[["Original file name",result.original_filename],["Validation date/time",result.validation_date.isoformat()+"Z"],["Rule pack used",result.rule_pack],["Overall status",result.summary.overall_status],["Error count",result.summary.errors],["Warning count",result.summary.warnings],["Info count",result.summary.info],["Sheets checked",result.summary.sheets_checked],["Rows checked",result.summary.rows_checked],["Rules executed",result.summary.rules_executed]]
    for r in rows: ws.append(r)
    ws2=wb.create_sheet("Sheet Summary"); ws2.append(["Sheet name","Rows checked","Errors","Warnings","Info","Status"])
    for s in result.sheet_summary: ws2.append([s.sheet_name,s.rows_checked,s.errors,s.warnings,s.info,s.status])
    ws3=wb.create_sheet("Issues"); ws3.append(ISSUE_FIELDS)
    for i in result.issues: ws3.append([getattr(i,f) for f in ISSUE_FIELDS])
    ws3.freeze_panes="A2"; ws3.auto_filter.ref=ws3.dimensions
    ws4=wb.create_sheet("Rule Summary"); ws4.append(["Rule ID","Rule name","Count","Severity"])
    for r in result.rule_summary: ws4.append([r.rule_id,r.rule_name,r.count,r.severity])
    ws5=wb.create_sheet("Metadata");
    for row in [["IFC Toolkit version","runtime"],["App environment",environment],["File size",file_size],["Processing duration",result.processing_duration_seconds],["Retention statement",result.metadata.get("retention_statement","")]]: ws5.append(row)
    for sheet in wb.worksheets:
        for cell in sheet[1]: cell.font=Font(bold=True)
        autosize(sheet)
    wb.save(path); return path

def write_marked_up_workbook(source: str|Path, result: CobieValidationResult, path: str|Path) -> Path:
    source=Path(source); path=Path(path); wb=load_workbook(source)
    fills={"Error":PatternFill("solid", fgColor="FFC7CE"),"Warning":PatternFill("solid", fgColor="FFEB9C"),"Info":PatternFill("solid", fgColor="D9EAF7")}
    for issue in result.issues:
        if issue.sheet_name in wb.sheetnames and issue.row_number>0 and issue.column_name:
            ws=wb[issue.sheet_name]; heads=[str(c.value or "").strip().lower() for c in ws[1]]
            if issue.column_name.lower() in heads:
                cell=ws.cell(issue.row_number, heads.index(issue.column_name.lower())+1)
                cell.fill=fills.get(issue.severity, fills["Info"])
                cell.comment=Comment(f"{issue.severity}: {issue.message}\nRecommendation: {issue.recommendation}", "IFC Toolkit")
    if "IFC Toolkit QA Report" in wb.sheetnames: del wb["IFC Toolkit QA Report"]
    ws=wb.create_sheet("IFC Toolkit QA Report"); ws.append(ISSUE_FIELDS)
    for i in result.issues: ws.append([getattr(i,f) for f in ISSUE_FIELDS])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions; autosize(ws); wb.save(path); return path
