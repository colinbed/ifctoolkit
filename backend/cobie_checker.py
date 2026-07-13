from __future__ import annotations
import re, time, uuid
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Set
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from backend.cobie_models import CobieIssue, CobieRuleSummary, CobieSheetSummary, CobieValidationResult, CobieValidationSummary
from backend.cobie_rules import CORE_SHEETS, DATE_FIELDS, FALLBACK_ALLOWED_VALUES, GENERIC_NAMES, NUMERIC_FIELDS, PLACEHOLDERS, TEXT_KEY_FIELDS, load_rulepack, load_sheet_config
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def _blank(v: Any) -> bool: return v is None or str(v).strip() == ""
def _txt(v: Any) -> str: return "" if v is None else str(v).strip()
def _is_date(v: Any) -> bool:
    if _blank(v): return True
    if isinstance(v, (datetime, date)): return True
    if isinstance(v, (int, float)): return v > 1
    s = _txt(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try: datetime.strptime(s[:19], fmt); return True
        except ValueError: pass
    return False
def _is_num(v: Any) -> bool:
    if _blank(v): return True
    if isinstance(v, (int, float)): return True
    try: float(str(v).strip()); return True
    except ValueError: return False

def validate_cobie_workbook(path: str | Path, rule_pack: str = "ifctoolkit-recommended", job_id: str = "", original_filename: str = "") -> CobieValidationResult:
    start = time.perf_counter(); cfg = load_sheet_config(); cfg_by_sheet={c["sheet_name"]:c for c in cfg}; pack=load_rulepack(rule_pack)
    wb = load_workbook(path, data_only=False)
    issues: List[CobieIssue]=[]; rules: Set[str]=set(); row_counts: Dict[str,int]=defaultdict(int); headers: Dict[str,List[str]]={}; records: Dict[str,List[tuple[int,Dict[str,Any]]]]={}
    def add(sev,sheet,row,col,rule,typ,msg,exp="",act="",rec="",rs=None,rc=None):
        rules.add(rule); cid=""; 
        if sheet in wb.sheetnames and col and row>0:
            hs=headers.get(sheet,[]); idx=next((i+1 for i,h in enumerate(hs) if h.lower()==col.lower()),0); cid=f"{sheet}!{get_column_letter(idx)}{row}" if idx else f"{sheet}!{row}"
        issues.append(CobieIssue(issue_id=f"COBIE-{len(issues)+1:05d}",severity=sev,sheet_name=sheet,row_number=row,column_name=col,cell_reference=cid,rule_id=rule,rule_type=typ,message=msg,expected_value=exp,actual_value=act,recommendation=rec,related_sheet=rs,related_column=rc))
    expected=set(cfg_by_sheet)
    for s,c in cfg_by_sheet.items():
        required = c.get("required") or s in CORE_SHEETS
        if s not in wb.sheetnames:
            add("Error" if required else "Warning",s,0,"",f"{s.upper()}_SHEET_MISSING","Workbook structure",f"COBie worksheet {s} is missing.",s,"Missing",f"Add the {s} worksheet using the COBie workbook structure.")
    for s in wb.sheetnames:
        if s not in expected: add("Warning",s,0,"","UNEXPECTED_WORKSHEET","Workbook structure",f"Worksheet {s} is not part of the configured COBie sheet list.","Configured COBie worksheet name",s,"Remove, rename, or confirm this project-specific worksheet.")
    for s in [x for x in wb.sheetnames if x in cfg_by_sheet]:
        ws=wb[s]; hs=[_txt(c.value) for c in ws[1]] if ws.max_row else [] ; headers[s]=hs
        seen=Counter([h.lower() for h in hs if h]);
        for i,h in enumerate(hs, start=1):
            if not h: add("Error",s,1,"",f"{s.upper()}_BLANK_HEADER","Workbook structure","Blank column heading found.","Column heading", "", "Populate or remove the blank heading.")
            elif seen[h.lower()]>1: add("Error",s,1,h,f"{s.upper()}_DUPLICATE_HEADER","Workbook structure",f"Duplicate column heading {h} found.","Unique heading",h,"Keep one heading and merge duplicated data.")
        expected_cols=cfg_by_sheet[s].get("expected_columns",[])
        for col in cfg_by_sheet[s].get("required_columns",[]):
            if col.lower() not in [h.lower() for h in hs]: add("Error",s,1,col,f"{s.upper()}_{col.upper()}_COLUMN_MISSING","Workbook structure",f"{s}.{col} column is required.",col,"Missing",f"Add the {col} column to {s}.")
        if pack.get("column_order_matters") and expected_cols and [h.lower() for h in hs[:len(expected_cols)]] != [h.lower() for h in expected_cols]:
            add("Error" if pack.get("strict_mode") else "Warning",s,1,"",f"{s.upper()}_COLUMN_ORDER","Workbook structure",f"{s} columns are not in configured COBie order.",", ".join(expected_cols),", ".join(hs),"Reorder columns to match the selected rule pack.")
        for h in hs:
            if h and expected_cols and h.lower() not in [e.lower() for e in expected_cols]: add("Info",s,1,h,f"{s.upper()}_UNEXPECTED_COLUMN","Workbook structure",f"{s}.{h} is an additional column.","Configured columns",h,"Confirm this project-specific field is intentional.")
        nonblank=0; recs=[]
        for r in range(2, ws.max_row+1):
            vals=[ws.cell(r,c).value for c in range(1, ws.max_column+1)]
            if all(_blank(v) for v in vals):
                if r < ws.max_row: add("Warning",s,r,"",f"{s.upper()}_BLANK_ROW","Data quality","Blank row found inside data range.","Populated COBie row","Blank row","Remove blank rows from the data range.")
                continue
            nonblank+=1; row_counts[s]+=1; rec={hs[i]: vals[i] for i in range(min(len(hs),len(vals))) if hs[i]}; recs.append((r,rec))
            for cidx,cell in enumerate(ws[r], start=1):
                if cell.data_type == "f": add("Warning",s,r,hs[cidx-1] if cidx<=len(hs) else "",f"{s.upper()}_FORMULA_CELL","Workbook structure","Formula cell found where a static COBie value is expected.","Static value",str(cell.value),"Replace formulas with reviewed static values before handover.")
        records[s]=recs
        if cfg_by_sheet[s].get("required") and nonblank==0: add("Error",s,2,"",f"{s.upper()}_EMPTY_REQUIRED_SHEET","Workbook structure",f"Required worksheet {s} has no data rows.","At least one data row","Empty",f"Populate {s} or adjust the rule pack.")
        for rng in ws.merged_cells.ranges:
            if rng.min_row >= 2 or rng.max_row >= 2: add("Error",s,rng.min_row,"",f"{s.upper()}_MERGED_CELLS","Workbook structure","Merged cells found in the COBie data area.","Unmerged cells",str(rng),"Unmerge cells and repeat values explicitly.")
    # indexes
    indexes: Dict[str,Set[str]]={}
    for s,recs in records.items():
        key = "Email" if s=="Contact" else "Name"
        indexes[s]={_txt(rec.get(key)) for _,rec in recs if not _blank(rec.get(key))}
    valid_sheets=set(cfg_by_sheet)
    allowed=dict(FALLBACK_ALLOWED_VALUES)
    if "PickLists" not in wb.sheetnames: add("Warning","PickLists",0,"","PICKLISTS_MISSING","PickLists","PickLists sheet is missing; built-in fallback allowed values were used.","PickLists worksheet","Missing","Add PickLists to enable project-specific allowed values.")
    # validations
    for s,recs in records.items():
        c=cfg_by_sheet.get(s,{}); req=c.get("required_columns",[])+(["ExternalIdentifier"] if pack.get("required_external_identifier") and s in {"Floor","Space","Type","Component"} else [])
        for row,rec in recs:
            for col in req:
                if col in rec and _blank(rec.get(col)): add("Error",s,row,col,f"{s.upper()}_{col.upper()}_REQUIRED","Required field",f"{s}.{col} is required and cannot be blank.",f"A valid {s}.{col} value",_txt(rec.get(col)),f"Populate {s}.{col} with the required COBie value.")
            if s=="Contact" and "Email" in rec and not _blank(rec.get("Email")) and not EMAIL_RE.match(_txt(rec.get("Email"))): add("Error",s,row,"Email","CONTACT_EMAIL_INVALID","Format","Contact.Email must look like a valid email address.","name@example.com",_txt(rec.get("Email")),"Correct the email address.")
            for col in DATE_FIELDS:
                if col in rec and not _is_date(rec.get(col)): add("Error",s,row,col,f"{s.upper()}_{col.upper()}_INVALID_DATE","Format",f"{s}.{col} must be a valid date.","Valid Excel or ISO date",_txt(rec.get(col)),"Use a valid date value.")
            for col in NUMERIC_FIELDS.get(s,[]):
                if col in rec and not _is_num(rec.get(col)): add("Warning",s,row,col,f"{s.upper()}_{col.upper()}_NUMERIC","Format",f"{s}.{col} should be numeric where provided.","Numeric value",_txt(rec.get(col)),"Use a numeric value rather than text.")
            for col,val in rec.items():
                text=_txt(val)
                if not text: continue
                if col in TEXT_KEY_FIELDS and text.lower() in PLACEHOLDERS and not pack.get("allow_placeholders"): add("Warning",s,row,col,f"{s.upper()}_{col.upper()}_PLACEHOLDER","Data quality",f"{s}.{col} contains a placeholder value.","Reviewed project value",text,"Replace placeholder values with confirmed information.")
                if col in TEXT_KEY_FIELDS and (text!=str(val) or "  " in text or "\n" in str(val)): add("Warning",s,row,col,f"{s.upper()}_{col.upper()}_TEXT_QUALITY","Data quality",f"{s}.{col} contains spacing or line-break quality issues.","Clean single-line text",str(val),"Trim whitespace and remove double spaces or line breaks.")
                if len(text)>int(pack.get("max_text_length",255)): add("Warning",s,row,col,f"{s.upper()}_{col.upper()}_TOO_LONG","Data quality",f"{s}.{col} is longer than the configured maximum.",str(pack.get("max_text_length",255)),str(len(text)),"Shorten the value or adjust the project rule pack.")
            for ref in c.get("reference_rules",[]):
                col=ref["column"]
                if col not in rec or _blank(rec.get(col)): continue
                targets=[x.strip() for x in _txt(rec.get(col)).split(",")] if ref.get("multi") else [_txt(rec.get(col))]
                for target in [t for t in targets if t]:
                    if target not in indexes.get(ref["sheet"],set()): add("Error",s,row,col,f"{s.upper()}_{col.upper()}_REFERENCE","Cross-sheet reference",f"{s}.{col} references a missing {ref['sheet']}.{ref['target_column']} value.",f"Existing {ref['sheet']}.{ref['target_column']}",target,f"Add the referenced {ref['sheet']} row or correct {s}.{col}.",ref["sheet"],ref["target_column"])
            if s in {"Attribute","Document","Coordinate","Issue"} and "SheetName" in rec and not _blank(rec.get("SheetName")):
                sh=_txt(rec.get("SheetName"));
                if sh not in valid_sheets: add("Error",s,row,"SheetName",f"{s.upper()}_SHEETNAME_REFERENCE","Cross-sheet reference",f"{s}.SheetName does not reference a valid COBie sheet.","Valid COBie sheet",sh,"Use a configured COBie worksheet name.")
                elif "RowName" in rec and not _blank(rec.get("RowName")) and _txt(rec.get("RowName")) not in indexes.get(sh,set()): add("Error",s,row,"RowName",f"{s.upper()}_ROWNAME_REFERENCE","Cross-sheet reference",f"{s}.RowName does not match a Name value on {sh}.",f"Existing {sh}.Name",_txt(rec.get("RowName")),"Correct RowName or add the referenced row.",sh,"Name")
            for key, vals in allowed.items():
                ss,cc=key.split(".")
                if ss==s and cc in rec and not _blank(rec.get(cc)) and _txt(rec.get(cc)).lower() not in vals: add("Warning",s,row,cc,f"{s.upper()}_{cc.upper()}_PICKLIST","PickLists",f"{s}.{cc} is not in the allowed value list.",", ".join(sorted(vals)),_txt(rec.get(cc)),"Use a configured PickLists value or update the project rule pack.")
        for col in c.get("unique_key_columns",[]):
            vals=[_txt(rec.get(col)) for _,rec in recs if not _blank(rec.get(col))]; counts=Counter(vals)
            for row,rec in recs:
                val=_txt(rec.get(col))
                if val and counts[val]>1: add("Error",s,row,col,f"{s.upper()}_{col.upper()}_DUPLICATE","Uniqueness",f"{s}.{col} must be unique.","Unique value",val,"Rename or merge duplicate records.")
    comp_types=[_txt(r.get("TypeName")) for _,r in records.get("Component",[]) if not _blank(r.get("TypeName"))]
    if len(set(comp_types))==1 and len(comp_types)>3: add("Warning","Component",0,"TypeName","COMPONENTS_SINGLE_TYPE","Data quality","All components are assigned to one TypeName.","Multiple appropriate TypeName values",comp_types[0],"Confirm this is expected or assign components to their correct asset types.")
    for row,rec in records.get("Component",[]):
        if _txt(rec.get("Name")).lower() in GENERIC_NAMES: add("Warning","Component",row,"Name","COMPONENT_GENERIC_NAME","Data quality","Component.Name appears to be generic.","Specific asset name",_txt(rec.get("Name")),"Use a specific component identifier.")
    # summaries
    by_sheet={s:CobieSheetSummary(sheet_name=s, rows_checked=row_counts.get(s,0)) for s in cfg_by_sheet if s in wb.sheetnames or s in CORE_SHEETS}
    by_rule={}
    for i in issues:
        ss=by_sheet.setdefault(i.sheet_name,CobieSheetSummary(sheet_name=i.sheet_name, rows_checked=row_counts.get(i.sheet_name,0)))
        if i.severity=="Error": ss.errors+=1
        elif i.severity=="Warning": ss.warnings+=1
        else: ss.info+=1
        ss.status="Fail" if ss.errors else ("Pass with warnings" if ss.warnings else ss.status)
        by_rule.setdefault(i.rule_id, {"count":0,"sev":i.severity,"name":i.rule_type})["count"]+=1
    errors=sum(1 for i in issues if i.severity=="Error"); warnings=sum(1 for i in issues if i.severity=="Warning"); infos=sum(1 for i in issues if i.severity=="Info")
    summary=CobieValidationSummary(overall_status="Fail" if errors else ("Pass with warnings" if warnings else "Pass"), total_issues=len(issues), errors=errors, warnings=warnings, info=infos, sheets_checked=len([s for s in wb.sheetnames if s in cfg_by_sheet]), rows_checked=sum(row_counts.values()), rules_executed=len(rules))
    return CobieValidationResult(job_id=job_id, original_filename=original_filename or Path(path).name, rule_pack=pack.get("id",rule_pack), processing_duration_seconds=round(time.perf_counter()-start,3), summary=summary, sheet_summary=list(by_sheet.values()), rule_summary=[CobieRuleSummary(rule_id=k,rule_name=v["name"],count=v["count"],severity=v["sev"]) for k,v in sorted(by_rule.items())], issues=issues, metadata={"retention_statement":"Uploaded COBie workbooks and generated reports are temporary session files and are not retained permanently."})
