import hashlib

import ifcopenshell
import pandas as pd
import pytest
from ifcopenshell.guid import new as new_guid
from openpyxl import load_workbook

from app import ExcelWritebackValidationError, extract_to_excel, update_ifc_from_excel


def _model(project_number="PN-SOURCE"):
    model = ifcopenshell.file(schema="IFC4")
    project = model.create_entity(
        "IfcProject", GlobalId=new_guid(), Name="Source project",
        Description="Project description", LongName=project_number,
    )
    site = model.create_entity("IfcSite", GlobalId=new_guid(), Name="Source site", Description="Site description")
    building = model.create_entity("IfcBuilding", GlobalId=new_guid(), Name="Source building")
    wall = model.create_entity("IfcWall", GlobalId=new_guid(), Name="Source wall", Description="Wall description")
    model.create_entity("IfcRelAggregates", GlobalId=new_guid(), RelatingObject=project, RelatedObjects=[site])
    model.create_entity("IfcRelAggregates", GlobalId=new_guid(), RelatingObject=site, RelatedObjects=[building])
    model.create_entity("IfcRelContainedInSpatialStructure", GlobalId=new_guid(), RelatingStructure=building, RelatedElements=[wall])
    return model


def _set_project_cell(path, field, value):
    workbook = load_workbook(path)
    sheet = workbook["ProjectData"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    project_row = next(row for row in range(2, sheet.max_row + 1) if sheet.cell(row, headers["DataType"]).value == "Project")
    sheet.cell(project_row, headers[field]).value = value
    workbook.save(path)
    workbook.close()


def _roundtrip(tmp_path, project_number="PN-SOURCE"):
    source = tmp_path / "source.ifc"
    workbook = tmp_path / "source.xlsx"
    output = tmp_path / "updated.ifc"
    _model(project_number).write(str(source))
    extract_to_excel(str(source), str(workbook))
    return source, workbook, output


def test_blank_project_number_preserves_source_and_applies_other_edit(tmp_path):
    source, workbook, output = _roundtrip(tmp_path)
    before = hashlib.sha256(source.read_bytes()).hexdigest()
    _set_project_cell(workbook, "ProjectNumber", None)
    _set_project_cell(workbook, "Name", "Edited project")

    result = update_ifc_from_excel(str(source), str(workbook), str(output))

    updated = ifcopenshell.open(str(output))
    projects = updated.by_type("IfcProject")
    assert len(projects) == 1
    assert projects[0].LongName == "PN-SOURCE"
    assert projects[0].Name == "Edited project"
    assert result["warnings"][0]["field"] == "ProjectNumber"
    assert hashlib.sha256(source.read_bytes()).hexdigest() == before
    assert updated.schema == ifcopenshell.open(str(source)).schema


def test_blank_project_number_can_remain_unset(tmp_path):
    source, workbook, output = _roundtrip(tmp_path, project_number=None)
    _set_project_cell(workbook, "ProjectNumber", "   ")

    result = update_ifc_from_excel(str(source), str(workbook), str(output))

    assert ifcopenshell.open(str(output)).by_type("IfcProject")[0].LongName is None
    assert "remain unset" in result["warnings"][0]["message"]


def test_changed_project_number_does_not_duplicate_project(tmp_path):
    source, workbook, output = _roundtrip(tmp_path)
    _set_project_cell(workbook, "ProjectNumber", "PN-EDITED")

    update_ifc_from_excel(str(source), str(workbook), str(output))

    projects = ifcopenshell.open(str(output)).by_type("IfcProject")
    assert len(projects) == 1
    assert projects[0].LongName == "PN-EDITED"


def test_blank_optional_metadata_is_preserved(tmp_path):
    source, workbook, output = _roundtrip(tmp_path)
    _set_project_cell(workbook, "Description", "  ")

    update_ifc_from_excel(str(source), str(workbook), str(output))

    assert ifcopenshell.open(str(output)).by_type("IfcProject")[0].Description == "Project description"


def test_invalid_entity_identifier_is_structured_error(tmp_path):
    source, workbook, output = _roundtrip(tmp_path)
    wb = load_workbook(workbook)
    sheet = wb["Elements"]
    guid_column = next(cell.column for cell in sheet[1] if cell.value == "GlobalId")
    sheet.cell(2, guid_column).value = "not-an-ifc-guid"
    wb.save(workbook)
    wb.close()

    with pytest.raises(ExcelWritebackValidationError) as caught:
        update_ifc_from_excel(str(source), str(workbook), str(output))

    error = caught.value.findings["errors"][0]
    assert error["sheet"] == "Elements"
    assert error["row"] == 2
    assert error["field"] == "GlobalId"
    assert not output.exists()
