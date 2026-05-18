from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

import app
from app import extract_to_excel, update_ifc_from_excel

IFC = """ISO-10303-21;
HEADER;
FILE_DESCRIPTION(('ViewDefinition [CoordinationView]'),'2;1');
FILE_NAME('x.ifc','2026-01-01T00:00:00',('a'),('o'),'EDMsix Version 602.28.129','Autodesk Civil 3D 2026','');
FILE_SCHEMA(('IFC2X3'));
ENDSEC;
DATA;
#1=IFCPROJECT('0J$X5dAxn7$Qf9f4f2x9d2',#2,'P',$,$,$,$,(#10),#20);
#2=IFCOWNERHISTORY($,$,.ADDED.,$,$,$,$,0);
#10=IFCGEOMETRICREPRESENTATIONCONTEXT($,'Model',3,1.E-05,#11,$);
#11=IFCAXIS2PLACEMENT3D(#12,$,$);
#12=IFCCARTESIANPOINT((0.,0.,0.));
#20=IFCUNITASSIGNMENT(());
#30=IFCBUILDINGSTOREY('3v6YQ6J8n0E8Qw4w9Nq9Jj',#2,'S',$,$,#31,$,.ELEMENT.,0.);
#31=IFCLOCALPLACEMENT($,#11);
ENDSEC;
END-ISO-10303-21;"""


def test_header_detection_edmsix_civil3d(tmp_path):
    p = tmp_path / "h.ifc"
    p.write_text(IFC, encoding="utf-8")
    meta = app.parse_ifc_header_metadata(str(p))
    assert meta["schema"] == "IFC2X3"
    assert meta["is_civil3d"] is True
    assert meta["is_edmsix"] is True


def test_storey_update_ifc2x3_no_has_associations_crash(tmp_path):
    src = tmp_path / "s.ifc"
    out = tmp_path / "o.ifc"
    src.write_text(IFC, encoding="utf-8")
    app.update_level(str(src), 30, {"name": "Updated", "cobie_floor": "L1"}, str(out))
    assert out.exists()


def test_export_openpyxl_valid(tmp_path):
    src = tmp_path / "a.ifc"
    xlsx = tmp_path / "a.xlsx"
    src.write_text(IFC, encoding="utf-8")
    extract_to_excel(str(src), str(xlsx), plan_payload={"include_sheets": ["ProjectData", "Elements", "COBieMapping", "Properties", "ChangeLog"]})
    wb = load_workbook(str(xlsx))
    assert "Elements" in wb.sheetnames
    wb.close()
