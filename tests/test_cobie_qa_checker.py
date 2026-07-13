from pathlib import Path
from openpyxl import Workbook
from backend.cobie_checker import validate_cobie_workbook

CORE = {
    "Contact":["Email","CreatedBy","CreatedOn","Company"],
    "Facility":["Name","CreatedBy","CreatedOn","ProjectName","SiteName","LinearUnits","AreaUnits","VolumeUnits","CurrencyUnit"],
    "Floor":["Name","CreatedBy","CreatedOn","Elevation","Height"],
    "Space":["Name","CreatedBy","CreatedOn","FloorName","GrossArea","NetArea"],
    "Type":["Name","CreatedBy","CreatedOn","Category","AssetType","WarrantyDurationParts"],
    "Component":["Name","CreatedBy","CreatedOn","TypeName","Space"],
    "System":["Name","CreatedBy","CreatedOn","ComponentNames"],
    "Attribute":["Name","CreatedBy","CreatedOn","SheetName","RowName","Value","Unit","Category"],
    "PickLists":["ListName","Value"],
}

def make_book(path: Path):
    wb=Workbook(); wb.remove(wb.active)
    rows={
        "Contact":[["alice@example.com","alice@example.com","2024-01-01","ACME"]],
        "Facility":[["HQ","alice@example.com","2024-01-01","Project","Site","meters","square meters","cubic meters","GBP"]],
        "Floor":[["L1","alice@example.com","2024-01-01",0,3]],
        "Space":[["Room 1","alice@example.com","2024-01-01","L1",10,9]],
        "Type":[["PumpType","alice@example.com","2024-01-01","Mechanical","fixed",12]],
        "Component":[["Pump-01","alice@example.com","2024-01-01","PumpType","Room 1"]],
        "System":[["Heating","alice@example.com","2024-01-01","Pump-01"]],
        "Attribute":[["Flow","alice@example.com","2024-01-01","Component","Pump-01","10","l/s","Performance"]],
        "PickLists":[["LinearUnits","meters"],["AreaUnits","square meters"]],
    }
    for s,heads in CORE.items():
        ws=wb.create_sheet(s); ws.append(heads)
        for r in rows.get(s,[]): ws.append(r)
    wb.save(path)

def rule_ids(result): return {i.rule_id for i in result.issues}

def test_valid_minimal_cobie_workbook_passes(tmp_path):
    p=tmp_path/'valid.xlsx'; make_book(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert result.summary.overall_status in {'Pass','Pass with warnings'}
    assert not [i for i in result.issues if i.severity=='Error']

def test_missing_required_sheet(tmp_path):
    p=tmp_path/'missing.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); del wb['Component']; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'COMPONENT_SHEET_MISSING' in rule_ids(result)
    assert result.summary.overall_status=='Fail'

def test_missing_required_column(tmp_path):
    p=tmp_path/'missing_col.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); ws=wb['Component']; ws.delete_cols(4); wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'COMPONENT_TYPENAME_COLUMN_MISSING' in rule_ids(result)

def test_blank_required_field_and_placeholder(tmp_path):
    p=tmp_path/'blank.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Component']['D2']='TBC'; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'COMPONENT_TYPENAME_PLACEHOLDER' in rule_ids(result)
    assert 'COMPONENT_TYPENAME_REFERENCE' in rule_ids(result)

def test_invalid_created_by_contact_reference(tmp_path):
    p=tmp_path/'bad_ref.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Space']['B2']='nobody@example.com'; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    issue=next(i for i in result.issues if i.rule_id=='SPACE_CREATEDBY_REFERENCE')
    assert issue.sheet_name=='Space' and issue.row_number==2 and issue.column_name=='CreatedBy'

def test_invalid_date(tmp_path):
    p=tmp_path/'bad_date.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Type']['C2']='not a date'; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'TYPE_CREATEDON_INVALID_DATE' in rule_ids(result)

def test_component_typename_references_missing_type(tmp_path):
    p=tmp_path/'missing_type.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Component']['D2']='MissingType'; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'COMPONENT_TYPENAME_REFERENCE' in rule_ids(result)

def test_space_floorname_references_missing_floor(tmp_path):
    p=tmp_path/'missing_floor.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Space']['D2']='L9'; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'SPACE_FLOORNAME_REFERENCE' in rule_ids(result)

def test_duplicate_component_name(tmp_path):
    p=tmp_path/'dup.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Component'].append(['Pump-01','alice@example.com','2024-01-01','PumpType','Room 1']); wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'COMPONENT_NAME_DUPLICATE' in rule_ids(result)

def test_picklists_missing_warns(tmp_path):
    p=tmp_path/'no_pick.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); del wb['PickLists']; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'PICKLISTS_MISSING' in rule_ids(result)

def test_picklist_invalid_value(tmp_path):
    p=tmp_path/'bad_pick.xlsx'; make_book(p); from openpyxl import load_workbook
    wb=load_workbook(p); wb['Facility']['F2']='parsecs'; wb.save(p)
    result=validate_cobie_workbook(p,'basic-cobie')
    assert 'FACILITY_LINEARUNITS_PICKLIST' in rule_ids(result)
