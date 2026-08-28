"""Regression guards for the memory-bounded IFC extraction architecture."""
from pathlib import Path

from openpyxl import load_workbook

from app import extract_to_excel


IFC = """ISO-10303-21;
HEADER;FILE_DESCRIPTION((''),'2;1');FILE_NAME('x.ifc','',(''),(''),'','','');FILE_SCHEMA(('IFC4'));ENDSEC;
DATA;
#1=IFCPROJECT('0J$X5dAxn7$Qf9f4f2x9d2',$,'Project',$,$,$,$,$,$);
#2=IFCBUILDINGELEMENTPROXY('1J$X5dAxn7$Qf9f4f2x9d2',$,'Asset',$,$,$,$,$,$);
#3=IFCPROPERTYSET('2J$X5dAxn7$Qf9f4f2x9d2',$,'Pset_Test',$,(#4));
#4=IFCPROPERTYSINGLEVALUE('Code',$,IFCLABEL('A-1'),$);
#5=IFCRELDEFINESBYPROPERTIES('3J$X5dAxn7$Qf9f4f2x9d2',$,$,$,(#2),#3);
ENDSEC;END-ISO-10303-21;
"""


def test_properties_are_streamed_and_counted(tmp_path: Path):
    source = tmp_path / "model.ifc"
    output = tmp_path / "model.xlsx"
    source.write_text(IFC, encoding="utf-8")
    result = extract_to_excel(str(source), str(output), plan_payload={"include_sheets": ["ProjectData", "Elements", "Properties"]})

    assert result["streaming"] is True
    assert result["counts"]["properties"] == 1
    assert "extract_properties" in result["timings_ms"]
    workbook = load_workbook(output, read_only=True)
    assert sum(1 for _ in workbook["Properties"].iter_rows()) == 2
    workbook.close()
